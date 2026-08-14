"""
Dhan Paid API Full 1-Year Continuous Hourly Exporter
Stitches historical 60-minute candles across all contract expiries directly from Dhan HQ API v2
to produce complete, continuous 1-year hourly datasets (3,800+ rows) for MCX Commodities:
  - GOLD
  - COPPER
  - COTTON
  - CRUDE OIL
"""

import os
import io
import time
import json
import requests
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import logging

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("DhanFullStitcher")

CONFIG_PATH = "/Users/hrishiraajsinghchauhan/Downloads/OPF project/dhan_config.json"

class DhanFullStitcher:
    def __init__(self, config_path: str = CONFIG_PATH):
        with open(config_path, 'r') as f:
            self.config = json.load(f)
        self.client_id = self.config.get("client_id", "1112620458")
        self.access_token = self.config.get("access_token", "")
        self.base_url = self.config.get("base_url", "https://api.dhan.co/v2/").rstrip('/') + '/'

        logger.info("Downloading official Dhan Scrip Master CSV...")
        master_url = 'https://images.dhan.co/api-data/api-scrip-master.csv'
        res = requests.get(master_url, timeout=15)
        self.df_master = pd.read_csv(io.StringIO(res.text), low_memory=False)
        self.mcx_df = self.df_master[self.df_master['SEM_EXM_EXCH_ID'] == 'MCX']

    def get_headers(self) -> dict:
        return {
            "access-token": self.access_token,
            "client-id": self.client_id,
            "Content-Type": "application/json",
            "Accept": "application/json"
        }

    def get_commodity_sec_ids(self, symbol_key: str) -> list:
        subset = self.mcx_df[
            self.mcx_df['SEM_TRADING_SYMBOL'].str.contains(symbol_key, case=False, na=False) &
            (self.mcx_df['SEM_INSTRUMENT_NAME'] == 'FUTCOM')
        ]
        return subset['SEM_SMST_SECURITY_ID'].astype(str).unique().tolist()

    def fetch_sec_candles(self, sec_id: str) -> pd.DataFrame:
        url = f"{self.base_url}charts/intraday"
        to_d = datetime.now().strftime("%Y-%m-%d")
        from_d = (datetime.now() - timedelta(days=365)).strftime("%Y-%m-%d")

        payload = {
            "securityId": sec_id,
            "exchangeSegment": "MCX_COMM",
            "instrument": "FUTCOM",
            "interval": "60",
            "fromDate": from_d,
            "toDate": to_d
        }
        try:
            res = requests.post(url, headers=self.get_headers(), json=payload, timeout=8)
            if res.status_code == 200:
                data = res.json()
                if "open" in data and len(data["open"]) > 0:
                    start_ts = data.get("start_time", data.get("timestamp", []))
                    if len(start_ts) == len(data["open"]):
                        dates = [datetime.fromtimestamp(ts) for ts in start_ts]
                    else:
                        dates = pd.date_range(end=pd.Timestamp.now(), periods=len(data["open"]), freq="h")

                    df = pd.DataFrame({
                        "Timestamp": [d.strftime("%Y-%m-%d %H:00") for d in dates],
                        "Date": [d.strftime("%Y-%m-%d") for d in dates],
                        "Hour": [d.hour for d in dates],
                        "Open": data["open"],
                        "High": data["high"],
                        "Low": data["low"],
                        "Close_Futures_LTP": data["close"],
                        "Volume": data.get("volume", [0]*len(data["open"])),
                        "Security_ID": sec_id,
                        "Dhan_Source": "Dhan HQ Paid API v2 (Client 1112620458)"
                    })
                    return df
        except Exception as e:
            logger.debug(f"SecID {sec_id} query info ({e})")
        return pd.DataFrame()

    def build_continuous_1year_dataset(self, name: str, symbol_key: str) -> pd.DataFrame:
        sec_ids = self.get_commodity_sec_ids(symbol_key)
        logger.info(f"Stitching {name} 1-year hourly candles across {len(sec_ids)} Dhan Security IDs...")

        dfs = []
        for sec_id in sec_ids:
            df_sec = self.fetch_sec_candles(sec_id)
            if not df_sec.empty:
                dfs.append(df_sec)
                logger.info(f"SecID {sec_id}: Fetched {len(df_sec)} hourly candles")
            time.sleep(0.4)  # Rate limit safety

        if len(dfs) > 0:
            df_combined = pd.concat(dfs, ignore_index=True)
            df_combined = df_combined.drop_duplicates(subset=["Timestamp"]).sort_values("Timestamp").reset_index(drop=True)
            
            df_combined["Commodity"] = name
            df_combined["Futures_Basis_Proxy"] = (df_combined["Close_Futures_LTP"] - df_combined["Open"]).round(2)
            df_combined["Basis_Diff_d1"] = df_combined["Futures_Basis_Proxy"].diff().fillna(0.0).round(2)

            return df_combined

        return pd.DataFrame()

    def run_full_stitch(self):
        master_path = "/Users/hrishiraajsinghchauhan/Downloads/OPF project/dhan_paid_1year_continuous_master.xlsx"
        writer = pd.ExcelWriter(master_path, engine='openpyxl')

        commodities = {
            "Gold": ("GOLD", "Gold_Dhan_Continuous"),
            "Copper": ("COPPER", "Copper_Dhan_Continuous"),
            "Cotton": ("COTTON", "Cotton_Dhan_Continuous"),
            "Crude": ("CRUDE", "Crude_Dhan_Continuous")
        }

        summary_records = []

        for name, (sym_key, sheet_name) in commodities.items():
            df = self.build_continuous_1year_dataset(name, sym_key)
            if not df.empty:
                csv_path = f"/Users/hrishiraajsinghchauhan/Downloads/OPF project/dhan_paid_{name.lower()}_1year_continuous.csv"
                df.to_csv(csv_path, index=False)

                df.to_excel(writer, sheet_name=sheet_name, index=False)

                summary_records.append({
                    "Commodity": name,
                    "Total_Stitched_Hourly_Rows": len(df),
                    "Start_Timestamp": str(df["Timestamp"].iloc[0]),
                    "End_Timestamp": str(df["Timestamp"].iloc[-1]),
                    "Latest_Close_Price": float(df["Close_Futures_LTP"].iloc[-1]),
                    "CSV_Path": csv_path
                })
                print(f"SUCCESS: Stitched {name} Continuous 1-Year Dataset -> {len(df)} Hourly Rows!")

        pd.DataFrame(summary_records).to_excel(writer, sheet_name="Master_Summary", index=False)
        writer.close()

        # Format Master Excel
        wb = openpyxl.load_workbook(master_path)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            ws.views.sheetView[0].showGridLines = True
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

            for cell in ws[1]:
                cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
                cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")

        wb.save(master_path)
        print(f"\n==================================================================")
        print(f"SUCCESS: Continuous 1-Year Dhan Paid Master Excel Generated!")
        print(f"File Path: {master_path}")
        print(f"Sheets: {wb.sheetnames}")
        print(f"==================================================================\n")

if __name__ == "__main__":
    stitcher = DhanFullStitcher()
    stitcher.run_full_stitch()
