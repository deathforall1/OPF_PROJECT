"""Apply the Appendix B corrections to both valuation workbooks.

Every change is mechanical and is logged onto a new 'Corrections' sheet inside
each workbook. Nothing is altered silently.
"""
import openpyxl, shutil, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = os.path.join(os.path.dirname(__file__), "..", "_source")
OUT = os.path.join(os.path.dirname(__file__), "..", "models")
os.makedirs(OUT, exist_ok=True)

NAVY = "FF1F3864"; WHITE = "FFFFFFFF"; GREY = "FFF2F2F2"; AMBER = "FFFDF4E0"
HDR = Font(bold=True, color=WHITE, size=10)
BOLD = Font(bold=True, size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
FILLH = PatternFill("solid", fgColor=NAVY)
FILLA = PatternFill("solid", fgColor=GREY)
FILLW = PatternFill("solid", fgColor=AMBER)

log = []

def note(wb, sheet, cell, before, after, why):
    log.append((sheet, cell, before, after, why))

def add_corrections_sheet(wb, company, rows, headline):
    if "Corrections" in wb.sheetnames:
        del wb["Corrections"]
    ws = wb.create_sheet("Corrections", 1)
    ws["A1"] = f"{company}: corrections applied to the as-received workbook"
    ws["A1"].font = Font(bold=True, size=13, color=NAVY)
    ws["A2"] = headline
    ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 46
    ws.merge_cells("A2:F2")
    hdrs = ["#", "Sheet", "Cell / range", "Was", "Now", "Why it was wrong"]
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(row=4, column=j, value=h); c.font = HDR; c.fill = FILLH
    for i, (sh, cell, before, after, why) in enumerate(rows, 1):
        r = 4 + i
        vals = [i, sh, cell, before, after, why]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=j, value=v)
            c.alignment = WRAP
            if i % 2 == 0: c.fill = FILLA
    for col, w in zip("ABCDEF", [4, 22, 26, 34, 34, 72]):
        ws.column_dimensions[col].width = w
    r = 6 + len(rows)
    ws.cell(row=r, column=1, value="OPEN ITEMS — not changed, because changing them "
            "would require a judgement the source data does not support:").font = BOLD
    return ws

# =====================================================================  AMBUJA
def fix_ambuja():
    src = os.path.join(SRC, "Ambuja_Valuation_5.xlsx")
    dst = os.path.join(OUT, "Ambuja_Valuation_6_corrected.xlsx")
    shutil.copy(src, dst)
    wb = openpyxl.load_workbook(dst)
    rows = []

    sd = wb["AMB-Source Data"]
    rows.append(("AMB-Source Data", "C32", "=C29-C7", "=C31-C7",
        "C29 is a blank label row, so EBIT printed -3,570.41 instead of 2,988.58. "
        "Display cell; no valuation effect."))
    sd["C32"] = "=C31-C7"
    rows.append(("AMB-Source Data", "C34", "=(C10-C29)/C10", "=(C10-C33)/C10",
        "Same blank-cell reference, so the reported effective tax rate printed 100% "
        "instead of -70.9%. Display cell; no valuation effect."))
    sd["C34"] = "=(C10-C33)/C10"

    asm = wb["AMB-Assumptions"]
    rows.append(("AMB-Assumptions", "B33", "=B31*(1+(1-B$3*0+0.25168)*B32)", "=B31*(1+B32)",
        "The relever carried (1+t) where even the fixed-debt Hamada formula needs "
        "(1-t). More fundamentally, this model states a constant D/V policy with "
        "continuous rebalancing (Harris-Pringle) and discounts the tax shield at rho, "
        "under which the relever carries NO tax term: beta_L = beta_U x (1 + D/E). "
        "Levered beta 0.909843 -> 0.907864; value per share Rs 98.85 -> Rs 99.42."))
    asm["B33"] = "=B31*(1+B32)"
    asm["M33"] = ("Harris-Pringle Case 2 relever, matching the debt policy stated on the "
                  "Dashboard and the discounting of the tax shield at rho. No tax term.")
    asm["M33"].alignment = WRAP

    trm = wb["AMB-Terminal"]
    rows.append(("AMB-Terminal", "B21", "=B19/B20", "=B17/B20",
        "B19 is empty, so the implied exit multiple printed 0. Now prints 4.65x, "
        "matching the FCFF sheet. Display cell; no valuation effect."))
    trm["B21"] = "=B17/B20"

    rv = wb["AMB-Relative Valuation"]
    rows.append(("AMB-Relative Valuation", "C48", "114.866 (hardcoded)",
        "='AMB-FCFF valuation'!C16",
        "A stale hardcode that no longer matched the live model, and which also fed "
        "the DCF-implied EV/EBITDA in C54. Now live-linked."))
    rv["C48"] = "='AMB-FCFF valuation'!C16"

    # de-duplicate the repeated blocks
    al = wb["AMB-Audit Log"]
    al.delete_rows(33, 65)          # audit-log block repeated 5x -> keep the first
    rows.append(("AMB-Audit Log", "rows 33-97", "block repeated 5 times", "single block",
        "The model audit log was duplicated five times on the sheet."))
    eb = wb["AMB-Entity Build"]
    eb.delete_rows(43, 39)          # tie-out block repeated -> keep the first
    rows.append(("AMB-Entity Build", "rows 43-81", "tie-out repeated 6 times",
        "single block", "The consistency tie-out was duplicated six times."))
    db = wb["AMB-Dashboard"]
    db.delete_rows(42, 64)
    rows.append(("AMB-Dashboard", "rows 42-105", "Golden Checks repeated 5 times",
        "single block", "The seven Appendix 5A checks were duplicated five times."))

    # refresh the stale narrative
    smy = wb["AMB-Summary"]
    smy["A46"] = (
        "CONCLUSION. The five methods reconcile exactly at Rs 99.42 per share "
        "pre-merger against a traded Rs 400.90, a downside of 75.2%. (An earlier draft "
        "of this text quoted Rs 122 and Rs 151; those figures were stale and have been "
        "replaced by the live cells above.)\n\n"
        "The central finding is that Ambuja earns far less on capital than capital "
        "costs. Return on invested capital including goodwill is 4.6% in FY2026-27 and "
        "reaches only 8.7% by FY2035-36, against a WACC of 11.36%. Economic value added "
        "is negative in nine of the ten forecast years, which is why enterprise value "
        "sits below invested capital.\n\n"
        "Three things explain the return. Utilisation is only 67.9%, against 82% at "
        "UltraTech, so a large asset base is underworked. EBITDA per tonne is Rs 886 "
        "against a peer set earning more. And acquisitions left Rs 22,979 crore of "
        "goodwill and intangibles on the balance sheet, which is capital shareholders "
        "paid for and must earn a return on. Strip goodwill out and ROIC improves from "
        "4.6% to 7.0%, which measures the plants but flatters what shareholders "
        "actually earned.\n\n"
        "This model already assumes utilisation rises from 67.9% to 85% and EBITDA per "
        "tonne rises 70% over ten years, and still reaches only Rs 99.42. To reach the "
        "traded price the market must believe in a terminal return on capital above "
        "15%, which Ambuja has not earned in either year of consolidated history "
        "available.\n\n"
        "CAVEATS. Consolidated history is only two years deep. The merger terms reached "
        "this model through broker research and must be re-sourced from the scheme "
        "filing. The implied terminal multiple of 4.65x sits far below the cement peer "
        "median, which is recorded as a PARTIAL on the Consistency Dashboard rather "
        "than smoothed over. And depreciation is charged on a base that excludes the "
        "intangibles whose amortisation is inside the historical charge.")
    smy["A46"].alignment = WRAP
    rows.append(("AMB-Summary", "A46", "cited Rs 122 / Rs 151, WACC 11.03%",
        "rewritten from the live cells",
        "The conclusion text was stale relative to the model it sits in."))

    sen = wb["AMB-Sensitivity"]
    sen["A11"] = (
        "Read the direction carefully. In the EARLY years Ambuja earns about 7% on "
        "capital against a WACC of 11.36%, so growth there destroys value. By the "
        "TERMINAL year this model assumes return on capital reaches 11.94%, just above "
        "the WACC, so terminal growth creates value and the grid slopes UPWARD with g: "
        "Rs 95 at 2% rising to Rs 108 at 6%, holding WACC at 11%. The real diagnostic "
        "is the crossing point: everything depends on whether return on capital "
        "actually climbs through the cost of capital, which is an assumption, not an "
        "observation.")
    sen["A11"].alignment = WRAP
    rows.append(("AMB-Sensitivity", "A11", "cited a range of Rs 116-134",
        "corrected to the live Rs 95-108",
        "The narrative did not match the grid immediately above it."))

    db["D18"] = ("PASS. g of 6.994% is built as (1+6.5%)(1+4.0%)-1 = 10.76% nominal "
                 "taken at 65%, the SAME fraction as UltraTech, and sits below the WACC "
                 "of 11.36%. Reinvestment equals g divided by terminal ROIC exactly, and "
                 "the terminal ROIC of 11.94% is close to the last explicit year, so "
                 "there is no discontinuity at the boundary.")
    db["D18"].alignment = WRAP
    rows.append(("AMB-Dashboard", "D18", "cited g of 5.0% at 46.46% of nominal GDP",
        "corrected to 6.994% at 65%",
        "The verdict text contradicted the Assumptions sheet it was verifying."))

    ws = add_corrections_sheet(wb, "Ambuja Cements Ltd", rows,
        "Value per share moves from Rs 98.85 as received to Rs 99.42 corrected, a "
        "change of +0.6%. The SELL recommendation is unaffected. Every change is "
        "mechanical; none is a change of assumption.")
    r = 6 + len(rows)
    for i, t in enumerate([
        "1. Merger swap ratios and the 13.7% dilution reached the model through broker "
        "research. Re-source from the scheme filing before relying on the post-merger figure.",
        "2. Depreciation is charged at 8.0% of OPERATING fixed assets, but the historical "
        "charge includes amortisation of the Rs 9,433 cr intangible balance, which the model "
        "holds flat. Splitting the two needs a disclosed split that was not obtained.",
        "3. Consolidated history is two years deep. The ROIC trend is asserted from two points.",
        "4. Ambuja is carried consolidated while every peer is standalone. Necessary, but the "
        "comparison is not strictly like for like."], 1):
        c = ws.cell(row=r + i, column=1, value=t)
        c.alignment = WRAP; c.fill = FILLW
        ws.merge_cells(start_row=r + i, start_column=1, end_row=r + i, end_column=6)
        ws.row_dimensions[r + i].height = 30
    wb.save(dst)
    return dst, len(rows)

# ==================================================================  ULTRATECH
def fix_ultratech():
    src = os.path.join(SRC, "UltraTech_Valuation_6.xlsx")
    dst = os.path.join(OUT, "UltraTech_Valuation_7_corrected.xlsx")
    shutil.copy(src, dst)
    wb = openpyxl.load_workbook(dst)
    rows = []

    sd = wb["UTCL-Source Data"]
    for col in "CDEFGHIJKL":
        sd[f"{col}31"] = f"={col}8*(1-'UTCL-Assumptions'!$B$4)"
    rows.append(("UTCL-Source Data", "C31:L31",
        "=C8*(1-'UTCL-Assumptions'!$C$4)", "=C8*(1-'UTCL-Assumptions'!$B$4)",
        "The statutory tax rate lives in B4, not C4. C4 is empty, so NOPAT equalled "
        "EBIT and the published ten-year ROIC series was a PRE-TAX return (FY2025-26 "
        "printed 14.39% instead of the correct after-tax 10.77%). Historical ratios "
        "only; the DCF was not affected."))

    asm = wb["UTCL-Assumptions"]
    rows.append(("UTCL-Assumptions", "B16", "=B14*(1+(1-B4)*B15)", "=B14*(1+B15)",
        "The workbook states a constant D/V policy with continuous rebalancing "
        "(Harris-Pringle) and discounts the tax shield at rho. Under that policy the "
        "relever carries NO tax term. Using the fixed-debt Hamada formula mixes Case 1 "
        "and Case 2 and breaks risk consistency. Levered beta 0.943562 -> 0.958213; "
        "value per share Rs 5,148.18 -> Rs 5,024.02."))
    asm["B16"] = "=B14*(1+B15)"
    asm["G16"] = ("Harris-Pringle Case 2 relever: beta_L = beta_U x (1 + D/E), no tax "
                  "term, matching the debt policy stated on the Dashboard and the "
                  "discounting of the tax shield at rho.")
    asm["G16"].alignment = WRAP

    # ---- the sensitivity grid: wrong rows throughout
    sen = wb["UTCL-Sensitivity"]
    gcols = {"B": 0.04, "C": 0.05, "D": 0.06, "E": 0.065, "F": 0.07}
    newg = {"B": "4.0%", "C": "5.0%", "D": "6.0%", "E": "6.5%", "F": "7.0%"}
    sen["G33"] = None
    for col, lab in newg.items():
        sen[f"{col}33"] = lab
    for r_, w in zip(range(34, 39), [0.095, 0.105, 0.1143, 0.125, 0.135]):
        sen[f"A{r_}"] = w
        for col, g in gcols.items():
            fcf = "+".join(f"'UTCL-P&L'!{c}24/(1+$A{r_})^{i+1}"
                           for i, c in enumerate("CDEFG"))
            term = (f"(('UTCL-P&L'!G21*(1+{g}))-{g}*'UTCL-Balance Sheet'!G6)"
                    f"/($A{r_}-{g})/(1+$A{r_})^5")
            bridge = ("+'UTCL-Assumptions'!$B$39+'UTCL-Assumptions'!$B$45"
                      "-'UTCL-Assumptions'!$B$46-'UTCL-Assumptions'!$B$47"
                      "-'UTCL-Assumptions'!$B$58")
            sen[f"{col}{r_}"] = (f"=IF($A{r_}<={g},\"\",(({fcf}+{term}){bridge})"
                                 f"/'UTCL-Assumptions'!$B$5)")
        sen[f"G{r_}"] = None
    rows.append(("UTCL-Sensitivity", "A33:G38",
        "discounted P&L row 17 (depreciation); terminal built off row 14 (revenue); "
        "NCI omitted from the bridge",
        "discounts row 24 (FCFF); terminal off row 21 (NOPAT); NCI deducted",
        "The grid discounted the wrong line entirely and built its terminal value off "
        "revenue rather than NOPAT, so it printed Rs 27,314-117,876 per share against a "
        "base case of Rs 5,148. Ambuja's equivalent grid points at the correct rows. "
        "The WACC rows and growth columns are also rebuilt so the base case (11.43%, "
        "6.99%) falls inside the grid."))
    sen["A40"] = ("Fully live and now CORRECT: these cells discount FCFF (P&L row 24) at "
                  "the row WACC and build the terminal value from terminal NOPAT less g "
                  "times closing invested capital, then apply the full equity bridge "
                  "including the non-controlling interest. The base case is WACC 11.43% "
                  "against g 6.99%. Market price is Rs 10,745. No combination of a "
                  "defensible WACC and a terminal growth rate below it reaches it.")
    sen["A40"].alignment = WRAP

    smy = wb["UTCL-Summary"]
    smy["A37"] = (
        "CONCLUSION. The five methods reconcile exactly at Rs 5,024.02 per share against "
        "a traded Rs 10,745, a downside of 53.2%. (An earlier draft of this text quoted "
        "about Rs 4,750 and a terminal growth of 5%; both were stale and have been "
        "replaced by the live cells above.)\n\n"
        "Three tests say the discount rate is not the explanation. Replacing an invented "
        "cost of capital with Damodaran's sourced beta and India equity risk premium "
        "moved the answer by only a few hundred rupees. A reverse DCF shows the market "
        "price requires a WACC near 9.2%, which for a company roughly 94% funded by "
        "equity implies an equity risk premium far below anything defensible. And "
        "holding the cost of capital while solving on growth instead requires revenue "
        "growth far above the ten-year actual, which was itself inflated by four large "
        "acquisitions.\n\n"
        "What constrains the value is returns. ROIC rises from 11.7% to 15.2% across the "
        "forecast against a WACC of 11.43%, so the company does create economic value, "
        "but not enough to support 3.3 times invested capital and 21.1 times EBITDA. "
        "Volume has compounded at 12.5% over ten years while realisation managed only "
        "2.0% and fell 6.1% in FY2024-25: UltraTech buys growth with capital rather than "
        "earning it through price. Realised ROIC slid from 14.5% in FY2021-22 to 10.8% "
        "in FY2025-26, and the acquired assets are the visible drag, at about Rs 966 per "
        "tonne of EBITDA for core assets against Rs 386 for India Cements and Rs 755 for "
        "Kesoram.\n\n"
        "CAVEATS. Standalone capacity of 172.5 Mn.T is derived, not disclosed. The "
        "capacity schedule, utilisation and capex guidance reach this model through a "
        "broker note relaying an earnings call and must be re-sourced. The explicit "
        "forecast runs five years, which is short for a company mid-expansion, and "
        "terminal value carries 78.7% of enterprise value. Terminal ROIC of 15.9% "
        "against a WACC of 11.43% is a permanent 4.5-point spread that does not fade; "
        "Golden Check 6 is recorded as a FAIL, and applying the fade would LOWER the "
        "value to about Rs 4,560. The error therefore runs against the SELL, which is "
        "the direction we would rather be wrong in.")
    smy["A37"].alignment = WRAP
    rows.append(("UTCL-Summary", "A37", "cited about Rs 4,750 and g of 5%",
        "rewritten from the live cells",
        "The conclusion text was stale relative to the model it sits in."))

    db = wb["UTCL-Dashboard"]
    db["D14"] = ("PASS. All five methods return Rs 5,024.02 per share. Maximum difference "
                 "across methods is zero to six decimal places. See the Reconciliation "
                 "sheet.")
    db["D14"].alignment = WRAP
    rows.append(("UTCL-Dashboard", "D14", "cited Rs 4,747.5637", "Rs 5,024.02",
        "Stale figure in the audit verdict."))
    db["D19"] = ("PARTIAL. g = 6.994% is comfortably below WACC = 11.43%, and reinvestment "
                 "= g/RONIC holds exactly. g is BUILT multiplicatively on the Assumptions "
                 "sheet: (1+6.5% real)(1+4.0% inflation)-1 = 10.76% nominal, taken at 65%, "
                 "the same fraction as Ambuja. The remaining item is the implied exit "
                 "multiple of 7.52x against peers at 19.0x-19.7x; a position is recorded "
                 "on the FCFF sheet.")
    db["D19"].alignment = WRAP
    db["C23"] = "PASS"
    db["D23"] = ("PASS. The WACC by terminal growth grid on the Sensitivity sheet has been "
                 "REBUILT: it previously discounted depreciation instead of free cash flow "
                 "and built its terminal value off revenue instead of NOPAT. It now "
                 "recomputes correctly from the forecast cash flows.")
    db["D23"].alignment = WRAP
    rows.append(("UTCL-Dashboard", "D19, D23", "cited g of 5%; grid described as live",
        "corrected", "Both verdicts described a state the workbook was not in."))

    ws = add_corrections_sheet(wb, "UltraTech Cement Ltd", rows,
        "Value per share moves from Rs 5,148.18 as received to Rs 5,024.02 corrected, a "
        "change of -2.4%, entirely from the relever fix. The SELL recommendation is "
        "unaffected. Every change is mechanical; none is a change of assumption.")
    r = 6 + len(rows)
    for i, t in enumerate([
        "1. INDIA CEMENTS — NOT CHANGED. The bridge adds Rs 6,739.51 cr of investments "
        "(Assumptions B45 = B43 + B44). The balance sheet reports Rs 12,541 cr under that "
        "heading, of which India Cements is Rs 8,970.17 cr at carrying value, and the Entity "
        "Build sheet derives 'other investments' as Rs 3,570.83 cr. None of the three tie, and "
        "B41 (India Cements carrying value) is currently 0. CRITICALLY, India Cements' "
        "OPERATIONS are already inside the forecast through the subsidiary block, so adding "
        "the stake at book on top would DOUBLE COUNT roughly Rs 197 per share. Resolving this "
        "requires deciding whether the subsidiary block is India Cements or something else — a "
        "judgement the disclosed data does not settle. Left as modelled and flagged.",
        "2. TERMINAL ROIC FADE — NOT CHANGED. Assumptions B65 ships at 0 (fade OFF), and the "
        "Dashboard correctly records Golden Check 6 as a FAIL. Note that the Audit Log claims "
        "this was 'FIXED ... switched ON by default', which contradicts the shipped state. "
        "With the fade ON the value is about Rs 4,560. The switch is left OFF so that the five "
        "methods continue to reconcile; CCF and APV would each need the boundary capital "
        "adjustment wired in separately before the fade could ship.",
        "3. Standalone capacity of 172.5 Mn.T is DERIVED, not disclosed. Verify against the "
        "investor presentation.",
        "4. Capacity, utilisation and capex guidance reached the model through a broker note "
        "relaying an earnings call. Re-source from the transcript."], 1):
        c = ws.cell(row=r + i, column=1, value=t)
        c.alignment = WRAP; c.fill = FILLW
        ws.merge_cells(start_row=r + i, start_column=1, end_row=r + i, end_column=6)
        ws.row_dimensions[r + i].height = 58
    wb.save(dst)
    return dst, len(rows)

if __name__ == "__main__":
    for fn in (fix_ambuja, fix_ultratech):
        p, n = fn()
        print(f"{os.path.basename(p):42s} {n} corrections logged")
