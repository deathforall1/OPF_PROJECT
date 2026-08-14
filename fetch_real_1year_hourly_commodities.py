"""
Authentic 1-Year Hourly Commodity Market Data Fetcher & Excel Exporter
Fetches real 1-year intraday 60-minute candles for Gold, Copper, Cotton, and Crude Oil futures
directly from financial market data feeds. Formats and exports individual CSV/Excel files
and a combined Master Excel Workbook.
"""

import numpy as np
import pandas as pd
import yfinance as yf
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import logging

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("RealCommoditiesFetcher")

TICKER_MAP = {
    "Gold": {"ticker": "GC=F", "unit": "USD/oz", "sheet": "Gold_Hourly", "csv": "real_gold_1year_hourly.csv", "excel": "real_gold_1year_hourly.xlsx"},
    "Copper": {"ticker": "HG=F", "unit": "USD/lb", "sheet": "Copper_Hourly", "csv": "real_copper_1year_hourly.csv", "excel": "real_copper_1year_hourly.xlsx"},
    "Cotton": {"ticker": "CT=F", "unit": "USD/lb", "sheet": "Cotton_Hourly", "csv": "real_cotton_1year_hourly.csv", "excel": "real_cotton_1year_hourly.xlsx"},
    "Crude": {"ticker": "CL=F", "unit": "USD/bbl", "sheet": "Crude_Hourly", "csv": "real_crude_1year_hourly.csv", "excel": "real_crude_1year_hourly.xlsx"}
}

def fetch_and_export_all_real_commodities():
    logger.info("Downloading authentic 1-year hourly futures market candles for Gold, Copper, Cotton, and Crude Oil...")

    master_excel_path = "/Users/hrishiraajsinghchauhan/Downloads/OPF project/real_commodities_1year_hourly_master.xlsx"
    writer = pd.ExcelWriter(master_excel_path, engine='openpyxl')

    summary_records = []

    for name, config in TICKER_MAP.items():
        ticker_symbol = config["ticker"]
        logger.info(f"Downloading 1-year hourly candles for {name} ({ticker_symbol})...")

        try:
            raw_df = yf.download(ticker_symbol, period="1y", interval="1h", progress=False)

            if raw_df.empty:
                raise ValueError(f"Downloaded empty data for {ticker_symbol}")

            # Handle MultiIndex columns if present
            if isinstance(raw_df.columns, pd.MultiIndex):
                raw_df = raw_df.xs(ticker_symbol, level=1, axis=1)

            df = pd.DataFrame(index=raw_df.index)
            df["Timestamp"] = [d.strftime("%Y-%m-%d %H:00") for d in raw_df.index]
            df["Date"] = [d.strftime("%Y-%m-%d") for d in raw_df.index]
            df["Hour"] = [d.hour for d in raw_df.index]
            df["Commodity"] = name
            df["Ticker"] = ticker_symbol
            df["Open"] = raw_df["Open"].round(3)
            df["High"] = raw_df["High"].round(3)
            df["Low"] = raw_df["Low"].round(3)
            df["Close_Futures_LTP"] = raw_df["Close"].round(3)
            df["Volume"] = raw_df["Volume"].fillna(0).astype(int)
            df["Hourly_Return_Pct"] = df["Close_Futures_LTP"].pct_change().fillna(0.0).round(4) * 100.0
            
            # Basis proxy (Close - Open hourly spread) & 1st differenced series
            df["Futures_Basis_Proxy"] = (df["Close_Futures_LTP"] - df["Open"]).round(3)
            df["Basis_Diff_d1"] = df["Futures_Basis_Proxy"].diff().fillna(0.0).round(3)
            df["Unit"] = config["unit"]
            df["Source"] = "Authentic Futures Market Feed (1-Year 60m Candles)"

            df = df.dropna(subset=["Close_Futures_LTP"])

            # Export individual CSV & Excel
            csv_path = f"/Users/hrishiraajsinghchauhan/Downloads/OPF project/{config['csv']}"
            excel_path = f"/Users/hrishiraajsinghchauhan/Downloads/OPF project/{config['excel']}"

            df.to_csv(csv_path, index=False)

            with pd.ExcelWriter(excel_path, engine='openpyxl') as ind_writer:
                df.to_excel(ind_writer, sheet_name=config["sheet"], index=False)
                format_sheet(ind_writer.sheets[config["sheet"]])

            # Append to Master Excel
            df.to_excel(writer, sheet_name=config["sheet"], index=False)

            summary_records.append({
                "Commodity": name,
                "Ticker": ticker_symbol,
                "Unit": config["unit"],
                "Total_Hourly_Rows": len(df),
                "Start_Timestamp": str(df["Timestamp"].iloc[0]),
                "End_Timestamp": str(df["Timestamp"].iloc[-1]),
                "Latest_Price": float(df["Close_Futures_LTP"].iloc[-1]),
                "CSV_File": config['csv'],
                "Excel_File": config['excel']
            })

            print(f"SUCCESS: Fetched & Exported {name} Real Hourly Data ({len(df)} rows)")

        except Exception as e:
            logger.error(f"Failed to fetch real data for {name} ({e})")

    # Add Summary Sheet to Master Excel
    pd.DataFrame(summary_records).to_excel(writer, sheet_name="Summary_Metadata", index=False)
    writer.close()

    # Format Master Excel
    wb = openpyxl.load_workbook(master_excel_path)
    for sheetname in wb.sheetnames:
        format_sheet(wb[sheetname])
    wb.save(master_excel_path)

    logger.info("Real Commodities Master Excel generation complete!")
    print(f"\n==================================================================")
    print(f"SUCCESS: Real 1-Year Hourly Commodities Master Excel Generated!")
    print(f"Master File Path: {master_excel_path}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"==================================================================\n")

def format_sheet(ws):
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")

    ws.views.sheetView[0].showGridLines = True
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

if __name__ == "__main__":
    fetch_and_export_all_real_commodities()
