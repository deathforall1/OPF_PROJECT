"""
Dhan Paid Data API 1-Year Historical Hourly Fetcher
Queries Dhan HQ API v2 Paid Historical/Intraday Endpoints:
  - POST https://api.dhan.co/v2/charts/historical
  - POST https://api.dhan.co/v2/charts/intraday
Fetches authentic 1-year hourly market candles directly from Dhan servers
for MCX Commodities (Gold, Copper, Cotton, Crude Oil).
"""

import os
import json
import requests
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import logging

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("DhanPaidHistoricalFetcher")

CONFIG_PATH = "/Users/hrishiraajsinghchauhan/Downloads/OPF project/dhan_config.json"

# Default MCX Security IDs on Dhan HQ API (can be updated in dhan_config.json)
DEFAULT_SECURITY_IDS = {
    "GOLD": {"security_id": "542381", "exchange": "MCX_COMM", "instrument": "FUTCOM", "sheet": "Gold_Hourly"},
    "COPPER": {"security_id": "542392", "exchange": "MCX_COMM", "instrument": "FUTCOM", "sheet": "Copper_Hourly"},
    "COTTON": {"security_id": "542405", "exchange": "MCX_COMM", "instrument": "FUTCOM", "sheet": "Cotton_Hourly"},
    "CRUDEOIL": {"security_id": "542370", "exchange": "MCX_COMM", "instrument": "FUTCOM", "sheet": "Crude_Hourly"}
}

class DhanPaidHistoricalFetcher:
    def __init__(self, config_path: str = CONFIG_PATH):
        self.config_path = config_path
        self.load_config()

    def load_config(self):
        if os.path.exists(self.config_path):
            with open(self.config_path, 'r') as f:
                self.config = json.load(f)
        else:
            self.config = {}

        self.client_id = self.config.get("client_id", "1112620458")
        self.access_token = self.config.get("access_token", "")
        self.base_url = self.config.get("base_url", "https://api.dhan.co/v2/").rstrip('/') + '/'

    def get_headers(self) -> dict:
        return {
            "access-token": self.access_token,
            "client-id": self.client_id,
            "Content-Type": "application/json",
            "Accept": "application/json"
        }

    def fetch_dhan_hourly_candles(self, symbol_key: str, sec_config: dict, from_date: str, to_date: str) -> pd.DataFrame:
        """
        Queries Dhan HQ API v2 Charts Intraday endpoint for 60-minute candles.
        Endpoint: POST https://api.dhan.co/v2/charts/intraday
        """
        url = f"{self.base_url}charts/intraday"
        payload = {
            "securityId": sec_config["security_id"],
            "exchangeSegment": sec_config["exchange"],
            "instrument": sec_config["instrument"],
            "instrumentType": sec_config["instrument"],
            "interval": "60",
            "fromDate": from_date,
            "toDate": to_date
        }

        logger.info(f"Querying Dhan Paid Data API for {symbol_key} ({from_date} to {to_date})...")

        try:
            res = requests.post(url, headers=self.get_headers(), json=payload, timeout=12)
            if res.status_code == 200:
                data = res.json()
                if "start_time" in data and len(data["start_time"]) > 0:
                    timestamps = [datetime.fromtimestamp(ts) for ts in data["start_time"]]
                    df = pd.DataFrame({
                        "Timestamp": [ts.strftime("%Y-%m-%d %H:00") for ts in timestamps],
                        "Date": [ts.strftime("%Y-%m-%d") for ts in timestamps],
                        "Hour": [ts.hour for ts in timestamps],
                        "Commodity": symbol_key,
                        "Open": data.get("open", []),
                        "High": data.get("high", []),
                        "Low": data.get("low", []),
                        "Close_Futures_LTP": data.get("close", []),
                        "Volume": data.get("volume", []),
                        "Open_Interest": data.get("open_interest", [0]*len(timestamps)),
                        "Dhan_Source": "Dhan HQ Paid Data API v2"
                    })
                    logger.info(f"Successfully fetched {len(df)} hourly candles for {symbol_key} from Dhan!")
                    return df
                else:
                    logger.warning(f"Dhan API returned empty candle payload for {symbol_key}: {data}")
            else:
                logger.error(f"Dhan API HTTP {res.status_code} Error: {res.text}")
        except Exception as e:
            logger.error(f"Failed to query Dhan API ({e})")

        return pd.DataFrame()

    def fetch_all_and_build_excel(self):
        if not self.access_token or self.access_token.startswith("YOUR_"):
            print("\n==================================================================")
            print("  DHAN PAID DATA API FETCH REQUIRES YOUR ACTIVE ACCESS TOKEN")
            print("==================================================================")
            print(f"Please open '{CONFIG_PATH}' and set your active access token:")
            print('{\n  "access_token": "YOUR_ACTUAL_DHAN_ACCESS_TOKEN_HERE"\n}')
            print("==================================================================\n")

        end_date = datetime.now().strftime("%Y-%m-%d")
        start_date = (datetime.now() - timedelta(days=365)).strftime("%Y-%m-%d")

        master_excel_path = "/Users/hrishiraajsinghchauhan/Downloads/OPF project/dhan_paid_master_commodities_hourly.xlsx"
        writer = pd.ExcelWriter(master_excel_path, engine='openpyxl')

        fetched_summary = []

        for symbol_key, sec_config in DEFAULT_SECURITY_IDS.items():
            df = self.fetch_dhan_hourly_candles(symbol_key, sec_config, start_date, end_date)

            if not df.empty:
                # Calculate basis diff series
                df["Futures_Basis"] = df["Close_Futures_LTP"] - df["Open"]
                df["Basis_Diff_d1"] = df["Futures_Basis"].diff().fillna(0.0)

                # Export to CSV
                csv_path = f"/Users/hrishiraajsinghchauhan/Downloads/OPF project/dhan_paid_{symbol_key.lower()}_hourly.csv"
                df.to_csv(csv_path, index=False)

                # Export sheet to Master Excel
                df.to_excel(writer, sheet_name=sec_config["sheet"], index=False)

                fetched_summary.append({
                    "Commodity": symbol_key,
                    "Security_ID": sec_config["security_id"],
                    "Status": "SUCCESS_DHAN_PAID_API",
                    "Hourly_Rows": len(df),
                    "From_Date": start_date,
                    "To_Date": end_date,
                    "CSV_Path": csv_path
                })
            else:
                fetched_summary.append({
                    "Commodity": symbol_key,
                    "Security_ID": sec_config["security_id"],
                    "Status": "REQUIRES_ACTIVE_DHAN_ACCESS_TOKEN",
                    "Hourly_Rows": 0,
                    "From_Date": start_date,
                    "To_Date": end_date
                })

        pd.DataFrame(fetched_summary).to_excel(writer, sheet_name="Dhan_Fetch_Summary", index=False)
        writer.close()

        # Format Excel Styling
        wb = openpyxl.load_workbook(master_excel_path)
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        align_center = Alignment(horizontal="center", vertical="center")

        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            ws.views.sheetView[0].showGridLines = True
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center

        wb.save(master_excel_path)
        print(f"\nSaved Dhan Paid Historical Master Excel Workbook -> {master_excel_path}")

if __name__ == "__main__":
    fetcher = DhanPaidHistoricalFetcher()
    fetcher.fetch_all_and_build_excel()
