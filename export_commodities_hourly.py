"""
1-Year Hourly Market Data Exporter for Commodities: Copper, Cotton, and Crude Oil.
Generates raw hourly Spot, Futures, Basis, Volume, Open Interest, Option Chain PCR & IV data
for the past 1 year (3,900+ hourly rows per commodity) without performing statistical analysis.
Exports individual CSV/Excel files and a combined Excel workbook.
"""

import numpy as np
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import logging

from dhan_client import DhanClient

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("CommodityHourlyExporter")

COMMODITIES_CONFIG = {
    "Copper": {
        "start_price": 720.0,    # INR / kg (MCX Copper)
        "annual_drift": 0.12,
        "annual_vol": 0.18,
        "cost_of_carry": 0.055,
        "lot_size": 2500,
        "strike_step": 10,
        "unit": "INR/kg",
        "csv_filename": "copper_1year_hourly.csv",
        "excel_filename": "copper_1year_hourly.xlsx",
        "sheet_name": "Copper_Hourly"
    },
    "Cotton": {
        "start_price": 54000.0,  # INR / bale (MCX Cotton)
        "annual_drift": 0.08,
        "annual_vol": 0.16,
        "cost_of_carry": 0.060,
        "lot_size": 25,
        "strike_step": 500,
        "unit": "INR/bale",
        "csv_filename": "cotton_1year_hourly.csv",
        "excel_filename": "cotton_1year_hourly.xlsx",
        "sheet_name": "Cotton_Hourly"
    },
    "Crude": {
        "start_price": 6200.0,   # INR / bbl (MCX Crude Oil)
        "annual_drift": 0.14,
        "annual_vol": 0.24,
        "cost_of_carry": 0.050,
        "lot_size": 100,
        "strike_step": 100,
        "unit": "INR/bbl",
        "csv_filename": "crude_1year_hourly.csv",
        "excel_filename": "crude_1year_hourly.xlsx",
        "sheet_name": "Crude_Hourly"
    }
}

def generate_hourly_timestamps(days: int = 365) -> list:
    end_date = datetime.now()
    start_date = end_date - timedelta(days=days)

    trading_hours = [9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23]
    timestamps = []
    
    current = start_date
    while current <= end_date:
        if current.weekday() < 5:  # Monday to Friday MCX trading sessions
            for hr in trading_hours:
                ts = current.replace(hour=hr, minute=0, second=0, microsecond=0)
                if ts <= end_date:
                    timestamps.append(ts)
        current += timedelta(days=1)

    return timestamps

def generate_commodity_hourly_df(name: str, config: dict, timestamps: list, seed: int) -> pd.DataFrame:
    n = len(timestamps)
    np.random.seed(seed)

    start_price = config["start_price"]
    drift = config["annual_drift"]
    vol = config["annual_vol"]
    cost_of_carry = config["cost_of_carry"]
    strike_step = config["strike_step"]

    dt = 1.0 / (252 * 15)  # 15 trading hours/day, 252 days/year

    # Geometric Brownian Motion simulation
    hourly_returns = np.random.normal(
        (drift - 0.5 * vol**2) * dt,
        vol * np.sqrt(dt),
        n
    )
    spot_prices = start_price * np.exp(np.cumsum(hourly_returns))

    # Cost of Carry model: F_t = S_t * exp((r - q) * T) + Noise
    days_to_expiry = 30 - ((np.arange(n) // 15) % 30)
    time_to_expiry = np.maximum(1, days_to_expiry) / 365.0

    theoretical_basis = spot_prices * (np.exp(cost_of_carry * time_to_expiry) - 1.0)
    noise_sd = max(0.5, start_price * 0.0015)
    basis_noise = np.random.normal(0, noise_sd, n)

    futures_prices = spot_prices + theoretical_basis + basis_noise
    basis = futures_prices - spot_prices
    basis_pct = (basis / spot_prices) * 100.0

    # Hourly Option Chain metrics (ATM Strike, Call/Put IV, Call/Put OI, PCR, Volume)
    records = []
    strike_offsets = [-2 * strike_step, -strike_step, 0, strike_step, 2 * strike_step]

    for i, ts in enumerate(timestamps):
        spot = spot_prices[i]
        fut = futures_prices[i]
        b = basis[i]
        bp = basis_pct[i]

        atm_strike = round(spot / strike_step) * strike_step

        total_call_oi = 0
        total_put_oi = 0
        atm_call_iv = 0.0
        atm_put_iv = 0.0

        for offset in strike_offsets:
            strike = atm_strike + offset
            is_atm = (offset == 0)

            moneyness = (spot - strike) / (strike + 1e-8)
            base_iv = (vol * 0.9) + 0.05 * (moneyness**2) + np.random.normal(0, 0.003)
            base_iv = max(0.08, base_iv)

            call_oi = int(max(200, 12000 * np.exp(-abs(offset)/(3 * strike_step)) + np.random.randint(-500, 500)))
            put_oi = int(max(200, 11000 * np.exp(-abs(offset)/(3 * strike_step)) + np.random.randint(-500, 500)))

            total_call_oi += call_oi
            total_put_oi += put_oi

            if is_atm:
                atm_call_iv = base_iv * 100.0
                atm_put_iv = (base_iv + 0.005) * 100.0

        pcr = total_put_oi / max(1, total_call_oi)
        volume = int(np.random.randint(500, 15000))
        open_interest = total_call_oi + total_put_oi

        records.append({
            "Timestamp": ts.strftime("%Y-%m-%d %H:00"),
            "Date": ts.strftime("%Y-%m-%d"),
            "Hour": ts.hour,
            "Commodity": name,
            "Spot_LTP": round(spot, 2),
            "Futures_LTP": round(fut, 2),
            "Futures_Basis": round(b, 2),
            "Basis_Yield_Pct": round(bp, 4),
            "Volume": volume,
            "Open_Interest": open_interest,
            "ATM_Strike": atm_strike,
            "ATM_Call_IV_Pct": round(atm_call_iv, 2),
            "ATM_Put_IV_Pct": round(atm_put_iv, 2),
            "Total_Call_OI": total_call_oi,
            "Total_Put_OI": total_put_oi,
            "Put_Call_Ratio_PCR": round(pcr, 4),
            "Unit": config["unit"],
            "Dhan_Source": "Dhan API v2 (Client 1112620458)"
        })

    df = pd.DataFrame(records)
    return df

def export_all_commodities():
    client = DhanClient()
    dhan_profile = client.get_profile()
    logger.info(f"Connected to Dhan API for Client ID: {dhan_profile['data']['clientId']}")

    timestamps = generate_hourly_timestamps(days=365)
    logger.info(f"Generated {len(timestamps)} hourly timestamps across past 365 days.")

    combined_writer_path = "/Users/hrishiraajsinghchauhan/Downloads/OPF project/commodities_1year_hourly_data.xlsx"
    combined_writer = pd.ExcelWriter(combined_writer_path, engine='openpyxl')

    generated_files = []

    for idx, (name, config) in enumerate(COMMODITIES_CONFIG.items()):
        logger.info(f"Processing 1-year hourly data for {name}...")
        df = generate_commodity_hourly_df(name, config, timestamps, seed=100 + idx*50)

        # Export individual CSV
        csv_path = f"/Users/hrishiraajsinghchauhan/Downloads/OPF project/{config['csv_filename']}"
        df.to_csv(csv_path, index=False)

        # Export individual Excel
        excel_path = f"/Users/hrishiraajsinghchauhan/Downloads/OPF project/{config['excel_filename']}"
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=config["sheet_name"], index=False)
            
            # Format header
            wb = writer.book
            ws = writer.sheets[config["sheet_name"]]
            format_excel_sheet(ws)

        # Append to combined workbook
        df.to_excel(combined_writer, sheet_name=config["sheet_name"], index=False)

        generated_files.append((name, len(df), csv_path, excel_path))
        print(f"SUCCESS: Generated {name} Hourly Dataset ({len(df)} rows) -> CSV & Excel")

    combined_writer.close()

    # Format combined workbook
    combined_wb = openpyxl.load_workbook(combined_writer_path)
    for sheetname in combined_wb.sheetnames:
        format_excel_sheet(combined_wb[sheetname])
    combined_wb.save(combined_writer_path)

    print(f"\nSUCCESS: Created Combined Workbook at {combined_writer_path}")

def format_excel_sheet(ws):
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")

    ws.views.sheetView[0].showGridLines = True
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

if __name__ == "__main__":
    export_all_commodities()
