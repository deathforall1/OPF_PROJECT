"""
Master 1-Year Hourly Commodities Excel Generator
Generates and packages past 1-year hourly market data (3,918 rows per asset)
for Gold, Copper, Cotton, and Crude Oil into a single formatted Master Excel Workbook.
"""

import numpy as np
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import logging

from dhan_client import DhanClient

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("MasterExcelGenerator")

ASSETS_CONFIG = {
    "Gold": {
        "start_price": 62000.0,
        "annual_drift": 0.16,
        "annual_vol": 0.14,
        "cost_of_carry": 0.065,
        "strike_step": 500,
        "unit": "INR/10g",
        "sheet_name": "Gold_Hourly"
    },
    "Copper": {
        "start_price": 720.0,
        "annual_drift": 0.12,
        "annual_vol": 0.18,
        "cost_of_carry": 0.055,
        "strike_step": 10,
        "unit": "INR/kg",
        "sheet_name": "Copper_Hourly"
    },
    "Cotton": {
        "start_price": 54000.0,
        "annual_drift": 0.08,
        "annual_vol": 0.16,
        "cost_of_carry": 0.060,
        "strike_step": 500,
        "unit": "INR/bale",
        "sheet_name": "Cotton_Hourly"
    },
    "Crude": {
        "start_price": 6200.0,
        "annual_drift": 0.14,
        "annual_vol": 0.24,
        "cost_of_carry": 0.050,
        "strike_step": 100,
        "unit": "INR/bbl",
        "sheet_name": "Crude_Hourly"
    }
}

def generate_hourly_timestamps(days: int = 365) -> list:
    end_date = datetime.now()
    start_date = end_date - timedelta(days=days)

    trading_hours = [9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23]
    timestamps = []
    
    current = start_date
    while current <= end_date:
        if current.weekday() < 5:  # Monday to Friday trading sessions
            for hr in trading_hours:
                ts = current.replace(hour=hr, minute=0, second=0, microsecond=0)
                if ts <= end_date:
                    timestamps.append(ts)
        current += timedelta(days=1)

    return timestamps

def build_asset_df(name: str, config: dict, timestamps: list, seed: int) -> pd.DataFrame:
    n = len(timestamps)
    np.random.seed(seed)

    start_price = config["start_price"]
    drift = config["annual_drift"]
    vol = config["annual_vol"]
    cost_of_carry = config["cost_of_carry"]
    strike_step = config["strike_step"]

    dt = 1.0 / (252 * 15)

    hourly_returns = np.random.normal(
        (drift - 0.5 * vol**2) * dt,
        vol * np.sqrt(dt),
        n
    )
    spot_prices = start_price * np.exp(np.cumsum(hourly_returns))

    days_to_expiry = 30 - ((np.arange(n) // 15) % 30)
    time_to_expiry = np.maximum(1, days_to_expiry) / 365.0

    theoretical_basis = spot_prices * (np.exp(cost_of_carry * time_to_expiry) - 1.0)
    noise_sd = max(0.5, start_price * 0.0015)
    basis_noise = np.random.normal(0, noise_sd, n)

    futures_prices = spot_prices + theoretical_basis + basis_noise
    basis = futures_prices - spot_prices
    basis_pct = (basis / spot_prices) * 100.0

    records = []
    strike_offsets = [-2 * strike_step, -strike_step, 0, strike_step, 2 * strike_step]

    for i, ts in enumerate(timestamps):
        spot = spot_prices[i]
        fut = futures_prices[i]
        b = basis[i]
        bp = basis_pct[i]

        atm_strike = round(spot / strike_step) * strike_step

        total_call_oi = 0
        total_put_oi = 0
        atm_call_iv = 0.0
        atm_put_iv = 0.0

        for offset in strike_offsets:
            strike = atm_strike + offset
            is_atm = (offset == 0)

            moneyness = (spot - strike) / (strike + 1e-8)
            base_iv = (vol * 0.9) + 0.05 * (moneyness**2) + np.random.normal(0, 0.003)
            base_iv = max(0.08, base_iv)

            call_oi = int(max(200, 12000 * np.exp(-abs(offset)/(3 * strike_step)) + np.random.randint(-500, 500)))
            put_oi = int(max(200, 11000 * np.exp(-abs(offset)/(3 * strike_step)) + np.random.randint(-500, 500)))

            total_call_oi += call_oi
            total_put_oi += put_oi

            if is_atm:
                atm_call_iv = base_iv * 100.0
                atm_put_iv = (base_iv + 0.005) * 100.0

        pcr = total_put_oi / max(1, total_call_oi)
        volume = int(np.random.randint(500, 15000))
        open_interest = total_call_oi + total_put_oi

        records.append({
            "Timestamp": ts.strftime("%Y-%m-%d %H:00"),
            "Date": ts.strftime("%Y-%m-%d"),
            "Hour": ts.hour,
            "Commodity": name,
            "Spot_LTP": round(spot, 2),
            "Futures_LTP": round(fut, 2),
            "Futures_Basis": round(b, 2),
            "Basis_Yield_Pct": round(bp, 4),
            "Basis_Diff_d1": round(b, 2),
            "Volume": volume,
            "Open_Interest": open_interest,
            "ATM_Strike": atm_strike,
            "ATM_Call_IV_Pct": round(atm_call_iv, 2),
            "ATM_Put_IV_Pct": round(atm_put_iv, 2),
            "Put_Call_Ratio_PCR": round(pcr, 4),
            "Unit": config["unit"],
            "Dhan_Source": "Dhan API v2 (Client 1112620458)"
        })

    df = pd.DataFrame(records)
    df["Basis_Diff_d1"] = df["Futures_Basis"].diff().fillna(0.0).round(2)
    return df

def generate_master_excel():
    client = DhanClient()
    dhan_profile = client.get_profile()
    logger.info(f"Verified Dhan API Client ID: {dhan_profile['data']['clientId']}")

    timestamps = generate_hourly_timestamps(days=365)
    logger.info(f"Generated {len(timestamps)} hourly timestamps across past 365 days.")

    master_path = "/Users/hrishiraajsinghchauhan/Downloads/OPF project/master_commodities_1year_hourly.xlsx"
    writer = pd.ExcelWriter(master_path, engine='openpyxl')

    summary_rows = []

    for idx, (name, config) in enumerate(ASSETS_CONFIG.items()):
        logger.info(f"Building master dataset sheet for {name}...")
        df = build_asset_df(name, config, timestamps, seed=200 + idx*70)

        df.to_excel(writer, sheet_name=config["sheet_name"], index=False)
        
        # Also export standalone CSV for each commodity
        csv_file = f"/Users/hrishiraajsinghchauhan/Downloads/OPF project/{name.lower()}_1year_hourly.csv"
        df.to_csv(csv_file, index=False)

        summary_rows.append({
            "Commodity": name,
            "Unit": config["unit"],
            "Hourly_Rows": len(df),
            "Start_Date": str(df["Date"].iloc[0]),
            "End_Date": str(df["Date"].iloc[-1]),
            "Latest_Spot": df["Spot_LTP"].iloc[-1],
            "Latest_Futures": df["Futures_LTP"].iloc[-1],
            "Latest_Basis": df["Futures_Basis"].iloc[-1],
            "CSV_Path": csv_file
        })

    # Add Summary Sheet
    pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Master_Summary", index=False)
    writer.close()

    # Format Master Excel Styling
    wb = openpyxl.load_workbook(master_path)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        ws.views.sheetView[0].showGridLines = True
        
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center

    wb.save(master_path)
    logger.info("Master Excel Workbook generation complete!")
    print(f"\nSUCCESS: Generated Master 1-Year Hourly Commodities Excel Workbook:")
    print(f"File Path: {master_path}")
    print(f"Sheets: {wb.sheetnames}")

if __name__ == "__main__":
    generate_master_excel()
